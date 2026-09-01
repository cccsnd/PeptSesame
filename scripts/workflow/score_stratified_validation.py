#!/usr/bin/env python3
"""Score-stratified validation (R1/R2 最关键要求): 检验 Layer 2 优先级分数是否
富集实验支持 (Wang et al. 2020 MS 验证) 的 sORF loci。

正集: 与 Wang 2020 Arabidopsis NCP 坐标重叠的 CDS-filter 后 sORF
负集: 长度 + 染色体匹配、不与任何 MS 肽重叠的 sORF (1:1)
指标: ROC-AUC / PR-AUC / top-1%/5%/10% 富集 / score 分布效应量
输出: results/08_benchmark/score_stratified_validation.tsv
"""
import csv
import random
import os, sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(os.environ.get("PEPTSESAME_ROOT", "."))
V5 = ROOT / "results"
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

OUT = V5 / "08_benchmark/score_stratified_validation.tsv"
RNG = random.Random(42)


def read_ms_peptides(xlsx):
    """读 Wang 2020 补充表 → [(chrom, start, end)]"""
    wb = load_workbook(xlsx, read_only=True)
    ws = wb.worksheets[0]
    recs = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        chrom = str(row[1]).strip()
        if chrom in ("Mt", "Pt", "None", ""):
            continue
        try:
            s, e = int(row[2]), int(row[3])
        except (TypeError, ValueError):
            continue
        recs.append((chrom, s, e))
    return recs


def overlaps(s1, e1, s2, e2):
    return s1 <= e2 and e1 >= s2


def main():
    # 1. MS 肽坐标
    ncp = read_ms_peptides(ROOT / "documents/maize_ara/mmc13.xlsx")   # NCP
    cp = read_ms_peptides(ROOT / "documents/maize_ara/mmc14.xlsx")    # CP
    print(f"MS 肽: NCP={len(ncp)} CP={len(cp)}")
    # 染色体索引
    ncp_idx = defaultdict(list)
    for chrom, s, e in ncp:
        ncp_idx[chrom].append((s, e))
    cp_idx = defaultdict(list)
    for chrom, s, e in cp:
        cp_idx[chrom].append((s, e))

    # 2. Arabidopsis scored sORF (CDS-filter 后)
    orfs = []  # (chrom, start, end, score, length_nt)
    with open(V5 / "02_layer2_scoring/Arabidopsis/scored_sorfs.tsv") as f:
        for r in csv.DictReader(f, delimiter="\t"):
            orfs.append((r["chrom"], int(r["start"]), int(r["end"]),
                         float(r["aggregated_score"]), int(r["length_nt"])))
    print(f"Arabidopsis CDS-filter 后 sORF: {len(orfs):,}")

    # 3. 正集 (NCP-overlapping) / 负集 (任何 MS 肽不重叠)
    def hit(chrom, s, e, idx):
        for a, b in idx.get(chrom, []):
            if overlaps(s, e, a, b):
                return True
        return False

    pos = [o for o in orfs if hit(o[0], o[1], o[2], ncp_idx)]
    # 负集候选: 不与 NCP 也不与 CP 重叠
    neg_pool = [o for o in orfs if not hit(o[0], o[1], o[2], ncp_idx) and not hit(o[0], o[1], o[2], cp_idx)]
    print(f"正集 (NCP-overlapping): {len(pos)}, 负池: {len(neg_pool):,}")

    # 4. 长度 + 染色体匹配抽样 1:1
    # 长度 bin (10-nt)
    def lbin(x):
        return (x // 10) * 10
    neg_by_key = defaultdict(list)
    for o in neg_pool:
        neg_by_key[(o[0], lbin(o[4]))].append(o)
    matched_neg = []
    for o in pos:
        key = (o[0], lbin(o[4]))
        cands = neg_by_key.get(key, [])
        if cands:
            matched_neg.append(RNG.choice(cands))
    print(f"匹配负集: {len(matched_neg)}")

    pos_scores = sorted([o[3] for o in pos], reverse=True)
    neg_scores = sorted([o[3] for o in matched_neg], reverse=True)

    # 5. ROC-AUC / PR-AUC
    try:
        from sklearn.metrics import roc_auc_score, average_precision_score
        y = [1] * len(pos) + [0] * len(matched_neg)
        sc = [o[3] for o in pos] + [o[3] for o in matched_neg]
        roc = roc_auc_score(y, sc)
        pr = average_precision_score(y, sc)
    except ImportError:
        roc = pr = float("nan")

    # 6. top-k 富集 (正集占比 vs 全体)
    n_pos, n_neg = len(pos), len(matched_neg)
    base_rate = n_pos / (n_pos + n_neg)
    def enrichment(k_frac):
        k = max(1, int(len(orfs) * k_frac))
        top = sorted(orfs, key=lambda o: o[3], reverse=True)[:k]
        top_pos = sum(1 for o in top if hit(o[0], o[1], o[2], ncp_idx))
        return top_pos / k

    rows = [
        ("n_positive_NCP_overlapping", n_pos),
        ("n_matched_negative", n_neg),
        ("roc_auc", round(roc, 4)),
        ("pr_auc", round(pr, 4)),
        ("base_positive_rate", round(base_rate, 4)),
        ("top1pct_positive_rate", round(enrichment(0.01), 4)),
        ("top5pct_positive_rate", round(enrichment(0.05), 4)),
        ("top10pct_positive_rate", round(enrichment(0.10), 4)),
        ("median_score_positive", round(sorted(pos_scores)[len(pos_scores)//2], 4)),
        ("median_score_negative", round(sorted(neg_scores)[len(neg_scores)//2], 4)),
    ]
    with open(OUT, "w", newline="") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow(["metric", "value"])
        w.writerows(rows)
    print(f"\n=== 结果 ===")
    for k, v in rows:
        print(f"  {k}: {v}")
    print(f"✅ {OUT}")


if __name__ == "__main__":
    main()
